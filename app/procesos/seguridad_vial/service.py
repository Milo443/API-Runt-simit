# ---------------------------------------------------------
# Autor: Camilo Calderon
# Github: @Milo443
# Nota: Desarrollado con amor por Milo
# Mensaje: usa el vibecoding y los agentes responsablemente.
# ---------------------------------------------------------
import io
import random
import time
import os
import json
from datetime import datetime
import openpyxl
import httpx
import asyncio
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from fastapi import HTTPException
from app.core.logging import logger
from app.core.config import settings
from ..runt.vehiculo.service import RuntService
from ..runt.ciudadano.service import CiudadanoService
from ..simit.service import SimitService
from .models import ExcelColumns
from app.core.websocket_manager import manager

class SeguridadVialService:
    @staticmethod
    def _copy_styles(source_cell, target_cell):
        """Helper to copy style from one cell to another."""
        if source_cell.has_style:
            target_cell.font = openpyxl.styles.Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                vertAlign=source_cell.font.vertAlign,
                underline=source_cell.font.underline,
                strike=source_cell.font.strike,
                color=source_cell.font.color
            )
            target_cell.border = openpyxl.styles.Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
            target_cell.fill = openpyxl.styles.PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = openpyxl.styles.Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                text_rotation=source_cell.alignment.text_rotation,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
                indent=source_cell.alignment.indent
            )

    @staticmethod
    def _get_task_path(client_id: str) -> str:
        return os.path.join(settings.STORAGE_MASS_RESULTS, f"{client_id}.json")

    @staticmethod
    def _get_result_path(client_id: str) -> str:
        return os.path.join(settings.STORAGE_MASS_RESULTS, f"{client_id}.xlsx")

    @staticmethod
    async def save_task_status(client_id: str, identificacion: str, status_data: dict):
        """Saves or updates the task status in a JSON file."""
        # Ensure the storage directory exists
        if not os.path.exists(settings.STORAGE_MASS_RESULTS):
            os.makedirs(settings.STORAGE_MASS_RESULTS, exist_ok=True)
            
        path = SeguridadVialService._get_task_path(client_id)
        # Preserve identificacion if not provided in status_data
        data = {**status_data, "identificacion": identificacion, "last_update": datetime.now().isoformat()}
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                try:
                    existing = json.load(f)
                    data = {**existing, **data}
                except: pass
        
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

    @staticmethod
    async def get_task_status(client_id: str, identificacion: str = None) -> dict:
        """Retrieves task status, ensuring it belongs to the correct identificacion."""
        path = SeguridadVialService._get_task_path(client_id)
        if not os.path.exists(path):
            return None
        
        with open(path, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
                if identificacion and data.get("identificacion") != identificacion:
                    return None
                return data
            except:
                return None

    @staticmethod
    async def cleanup_old_tasks():
        """Deletes tasks and results older than 24 hours."""
        if not os.path.exists(settings.STORAGE_MASS_RESULTS):
            return
        
        now = time.time()
        for filename in os.listdir(settings.STORAGE_MASS_RESULTS):
            path = os.path.join(settings.STORAGE_MASS_RESULTS, filename)
            if os.path.getmtime(path) < (now - 24 * 3600):
                try:
                    os.remove(path)
                    logger.info(f"Cleaned up old file: {filename}")
                except Exception as e:
                    logger.error(f"Error cleaning up {filename}: {e}")

    @staticmethod
    async def procesar_excel(file_content: bytes, es_conductor_laboral: str, rodamiento: str, client_id: str = None, identificacion: str = None) -> bytes:
        """
        Optimized report generation using concurrency, connection pooling, and fallback/retry mechanism.
        """
        try:
            start_time = time.time()
            
            # 1. Load the input file
            try:
                in_wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                if "ACTIVOS" not in in_wb.sheetnames:
                    raise HTTPException(status_code=400, detail="El archivo no tiene la hoja 'ACTIVOS'")
                in_ws = in_wb["ACTIVOS"]
            except Exception as e:
                logger.error(f"Error loading input Excel: {e}")
                raise HTTPException(status_code=400, detail=f"Error cargando el archivo Excel: {str(e)}")

            # 2. Setup output Workbook and Mapping
            current_dir = os.path.dirname(os.path.abspath(__file__))
            blueprint_path = os.path.join(current_dir, "blueprint", "blueprint_seguridad_vial.xlsx")
            mapping_path = os.path.join(current_dir, "mapping.json")
            
            try:
                out_wb = openpyxl.load_workbook(blueprint_path, keep_links=False)
                out_ws = out_wb["ACTIVOS"]
                with open(mapping_path, "r", encoding="utf-8") as f:
                    mapping = json.load(f)
                
                # Clear template row 10 values but keep styles
                for col in range(1, out_ws.max_column + 1):
                    out_ws.cell(row=10, column=col).value = None
            except Exception as e:
                logger.error(f"Error initializing Workbook or Mapping: {e}")
                raise HTTPException(status_code=500, detail="No se pudo preparar el reporte de salida")

            # 3. Enhanced Helpers
            def parse_date(date_str):
                if not date_str: return datetime.min
                date_str = str(date_str).strip()
                base_date = date_str.split(' ')[0] if ' ' in date_str else date_str
                try:
                    if "T" in base_date: return datetime.strptime(base_date[:10], "%Y-%m-%d")
                    return datetime.strptime(base_date, "%d/%m/%Y")
                except:
                    try: return datetime.strptime(base_date, "%Y-%m-%d")
                    except: return datetime.min

            def format_date(date_str):
                if not date_str: return ""
                dt = parse_date(date_str)
                return dt.strftime("%d/%m/%Y") if dt != datetime.min else str(date_str)

            def get_by_path(obj, path):
                if not obj: return None
                for part in path.split('.'):
                    if isinstance(obj, dict): obj = obj.get(part)
                    elif isinstance(obj, list) and part.isdigit():
                        idx = int(part)
                        obj = obj[idx] if idx < len(obj) else None
                    else: return None
                return obj

            def apply_mapping_logic(group_name, data_source, target_cell_dict, extra=None):
                """Returns values to be written instead of writing directly to worksheet."""
                group = mapping.get(group_name, {})
                for key, config in group.items():
                    col = config.get("column")
                    path = config.get("path")
                    strategy = config.get("strategy", "direct")
                    fmt = config.get("format")
                    
                    val = None
                    is_normativa = False
                    
                    if strategy == "direct":
                        val = get_by_path(data_source, path)
                    elif strategy == "latest":
                        base_path = path.split('.')[0]
                        items = data_source.get(base_path, [])
                        sort_path = config.get("sort_path")
                        if items and isinstance(items, list) and sort_path:
                            attr = path.split('.')[-1]
                            valid_items = [i for i in items if isinstance(i, dict) and get_by_path(i, sort_path)]
                            if valid_items:
                                latest_item = max(valid_items, key=lambda x: parse_date(get_by_path(x, sort_path)))
                                val = get_by_path(latest_item, attr)
                    elif strategy == "join":
                        base_path = path.split('.')[0]
                        items = data_source.get(base_path, [])
                        if items and isinstance(items, list):
                            attr = path.split('.')[-1]
                            vals = sorted(list(set([str(i.get(attr, "")) for i in items if isinstance(i, dict) and i.get(attr)])))
                            val = ", ".join(vals)
                    elif strategy == "join_dates":
                        base_path = path.split('.')[0]
                        items = data_source.get(base_path, [])
                        if items and isinstance(items, list):
                            attr = path.split('.')[-1]
                            date_attr = config.get("date_path", "fechaVencimiento")
                            res = []
                            for i in items:
                                if not isinstance(i, dict): continue
                                status = str(i.get("estadoLicenciaHeader") or i.get("estadoLicencia") or "").strip().upper()
                                if status != "ACTIVA": continue
                                c_val = str(i.get(attr, "")).strip()
                                d_val = i.get(date_attr)
                                if c_val:
                                    d_fmt = format_date(d_val)
                                    res.append(f"{c_val} ({d_fmt})" if d_fmt else c_val)
                            val = ", ".join(res)
                    elif strategy == "newline":
                        base_parts = path.split('.')
                        base_path = base_parts[0]
                        items = data_source.get(base_path, [])
                        if items and isinstance(items, list):
                            attr = ".".join(base_parts[1:])
                            prefix = config.get("prefix", "")
                            fmt_i = config.get("format")
                            res = []
                            for i in items:
                                if not isinstance(i, dict): continue
                                val_i = get_by_path(i, attr) if attr else i
                                if val_i is not None and str(val_i).strip():
                                    if fmt_i == "date": val_i = format_date(val_i)
                                    res.append(f"{prefix}{val_i}")
                            val = "\n".join(res)
                    elif strategy == "match_category":
                        base_path = path.split('.')[0]
                        items = data_source.get(base_path, [])
                        cat_labor = extra.get("categoria_labor") if extra else None
                        if cat_labor and items and isinstance(items, list):
                            import re
                            target_cats = set(re.split(r'[\|, \s]+', str(cat_labor).lower().strip()))
                            target_cats = {c for c in target_cats if c}
                            attr = path.split('.')[-1]
                            matched_lics = []
                            for lic in items:
                                if not isinstance(lic, dict): continue
                                lic_cat = str(lic.get("categoria", "")).lower().strip()
                                if lic_cat in target_cats: matched_lics.append(lic)
                            if matched_lics:
                                latest_lic = max(matched_lics, key=lambda x: parse_date(x.get("fechaVencimiento") or x.get("fechaVencimientoLicencia")))
                                val = latest_lic.get(attr)
                    elif strategy == "rtm_logic":
                        base_path = path.split('.')[0]
                        items = data_source.get(base_path, [])
                        attr = path.split('.')[-1]
                        if items and isinstance(items, list):
                            latest_item = max(items, key=lambda x: parse_date(x.get(config.get("sort_path"))))
                            val = latest_item.get(attr)
                        else:
                            clase = str(get_by_path(data_source, "informacion_general.clase") or "").strip().lower()
                            fecha_mat = str(get_by_path(data_source, "informacion_general.fechaRegistro") or "").strip()
                            if clase and fecha_mat:
                                motos = ["motocicleta", "motocarro", "mototriciclo", "cuatrimoto", "ciclomotor"]
                                carros = ["automóvil", "camioneta", "campero", "microbús", "buseta", "bus", "camión", "tractocamión", "volqueta"]
                                reg_dt = parse_date(fecha_mat)
                                if reg_dt != datetime.min:
                                    years_to_add = 0
                                    if clase in motos: years_to_add = 2
                                    elif clase in carros: years_to_add = 5
                                    if years_to_add:
                                        try: expiration_dt = reg_dt.replace(year=reg_dt.year + years_to_add)
                                        except ValueError: expiration_dt = reg_dt.replace(year=reg_dt.year + years_to_add, day=28)
                                        val = expiration_dt if attr == "fechaVencimientoRvt" else ("Vigente" if datetime.now() < expiration_dt else "No vigente")
                                        if val == "Vigente" or attr == "fechaVencimientoRvt": is_normativa = True

                    if fmt == "date": val = format_date(val)
                    if val is not None:
                        target_cell_dict[col] = {"value": val, "is_normativa": is_normativa, "is_newline": (strategy == "newline" or "\n" in str(val))}

            # 4. Filtering and Preconditioning
            all_valid_indices = []
            filtered_indices = []
            for r in range(10, in_ws.max_row + 1):
                c_id = str(in_ws.cell(row=r, column=ExcelColumns.CEDULA).value or "").strip()
                if c_id:
                    all_valid_indices.append(r)
                    v_cond = str(in_ws.cell(row=r, column=ExcelColumns.ES_CONDUCTOR_LABORAL).value or "").strip().lower()
                    v_rod = str(in_ws.cell(row=r, column=ExcelColumns.RODAMIENTO).value or "").strip().lower()
                    if v_cond == es_conductor_laboral.lower() and v_rod == rodamiento.lower():
                        filtered_indices.append(r)
            
            total_activos = len(all_valid_indices)
            total_filtrados = len(filtered_indices)
            
            logger.info(f"--------------------------------------------------")
            logger.info(f"INICIANDO PROCESO MASIVO: {client_id}")
            logger.info(f"Total filas con documento: {total_activos}")
            logger.info(f"Filas a procesar (según filtros): {total_filtrados}")
            logger.info(f"--------------------------------------------------")
            
            if client_id:
                status_update = {
                    "total": total_activos, 
                    "total_filtrados": total_filtrados,
                    "status": "iniciado", 
                    "progress": 0,
                    "identificacion": identificacion
                }
                await SeguridadVialService.save_task_status(client_id, identificacion, status_update)
                await manager.send_personal_message(status_update, client_id)

            # 5. Concurrent Processing Engine
            semaphore = asyncio.Semaphore(settings.BATCH_SIZE)
            cache_veh = {}
            cache_cit = {}
            cache_sim = {}
            results = {} # Map original_idx -> result_data
            
            # Progress & Status tracking
            processed_count = 0
            exitosos_count = 0
            fallidos_count = 0
            live_results_map = {} # my_index -> {status, cedula, nombre, index}
            counter_lock = asyncio.Lock()
            
            # Track current row index to send "procesando" status
            row_id_counter = 0
            row_id_lock = asyncio.Lock()

            async def run_with_retry(func, *args, max_retries=3, **kwargs):
                """Internal helper to execute a service call with a fallback (retry) mechanism."""
                last_error = None
                for attempt in range(1, max_retries + 1):
                    try:
                        res = await func(*args, **kwargs)
                        if res.get("exito"):
                            return res
                        last_error = res.get("mensaje") or res.get("error")
                        if attempt < max_retries:
                            delay = 1.0 + random.uniform(0.5, 1.5) # Dynamic jitter (1.5s to 2.5s)
                            logger.warning(f"Fallback Attempt {attempt}/{max_retries} for {func.__name__}: {last_error}. Retrying in {round(delay, 1)}s...")
                            await asyncio.sleep(delay)
                    except Exception as e:
                        last_error = str(e)
                        if attempt < max_retries:
                            delay = 1.0 + random.uniform(0.5, 1.5)
                            logger.error(f"Exc in Fallback Attempt {attempt}/{max_retries} for {func.__name__}: {last_error}. Retrying in {round(delay, 1)}s...")
                            await asyncio.sleep(delay)
                return {"exito": False, "error": last_error}

            async def query_row(row_idx, client):
                nonlocal processed_count, row_id_counter, exitosos_count, fallidos_count
                
                # Extract row basics
                cedula = str(in_ws.cell(row=row_idx, column=ExcelColumns.CEDULA).value or "").strip()
                nombre = str(in_ws.cell(row=row_idx, column=ExcelColumns.NOMBRE).value or "").strip()
                placa = str(in_ws.cell(row=row_idx, column=ExcelColumns.PLACA).value or "").strip()
                cat_labor = str(in_ws.cell(row=row_idx, column=ExcelColumns.LICENCIA_LABOR_CATEGORIA).value or "").strip()
                es_prop = str(in_ws.cell(row=row_idx, column=ExcelColumns.ES_PROPIETARIO).value or "").strip().lower()
                ident_prop = str(in_ws.cell(row=row_idx, column=ExcelColumns.IDENTIFICACION_PROPIETARIO).value or "").strip()
                doc_v = ident_prop if es_prop == "no" and ident_prop else cedula

                # USER FIX: Decide if we should consult or skip (traceability)
                should_consult = row_idx in filtered_indices
                row_result = {"status": "FILA_COMPLETADA" if should_consult else "No consultado", "cells": {}, "cedula": cedula, "nombre": nombre}
                my_index = -1

                if should_consult:
                    # Initially show row as 'esperando' until it enters semaphore
                    row_result["status"] = "esperando"
                    # Track this specific row's assigned index in the live results list
                    async with row_id_lock:
                        my_index = row_id_counter
                        row_id_counter += 1
                    
                    # Initial jitter per row to desynchronize burst started within the semaphore
                    initial_jitter = random.uniform(0.1, 1.2)
                    await asyncio.sleep(initial_jitter)

                    # Send initial "procesando" update
                    if client_id:
                        msg = {
                            "index": my_index, "total": total_activos, "total_filtrados": total_filtrados,
                            "cedula": cedula, "nombre": nombre, "status": "procesando"
                        }
                        async with counter_lock:
                            live_results_map[my_index] = {"status": "procesando", "cedula": cedula, "nombre": nombre, "index": my_index}
                        await manager.send_personal_message(msg, client_id)

                    async with semaphore:
                        try:
                            # Parallel Queries with Fallback/Retry Mechanism
                            tasks = []
                            
                            # Task A: RUNT Vehículo (with cache and internal retry)
                            key_veh = f"{placa}_{doc_v}"
                            if placa and doc_v:
                                if key_veh in cache_veh: tasks.append(asyncio.sleep(0, result=cache_veh[key_veh]))
                                else: tasks.append(run_with_retry(RuntService.consultar_vehiculo, placa, doc_v, client))
                            else: tasks.append(asyncio.sleep(0, result={"exito": False}))

                            # Task B: RUNT Ciudadano (with cache and internal retry)
                            if cedula in cache_cit: tasks.append(asyncio.sleep(0, result=cache_cit[cedula]))
                            else: tasks.append(run_with_retry(CiudadanoService.consultar_ciudadano, "C", cedula, client))

                            # Task C: SIMIT (with cache and internal retry)
                            if cedula in cache_sim: tasks.append(asyncio.sleep(0, result=cache_sim[cedula]))
                            else: tasks.append(run_with_retry(SimitService.consultar_ciudadano, cedula, client))

                            responses = await asyncio.gather(*tasks, return_exceptions=True)
                            
                            r_veh = responses[0] if not isinstance(responses[0], Exception) else {"exito": False, "error": str(responses[0])}
                            r_cit = responses[1] if not isinstance(responses[1], Exception) else {"exito": False, "error": str(responses[1])}
                            r_sim = responses[2] if not isinstance(responses[2], Exception) else {"exito": False, "error": str(responses[2])}

                            # Cache results
                            if r_veh.get("exito"): cache_veh[key_veh] = r_veh
                            if r_cit.get("exito"): cache_cit[cedula] = r_cit
                            if r_sim.get("exito"): cache_sim[cedula] = r_sim

                            # Map results to our cell structure
                            cells = row_result["cells"]
                            # RUNT Veh
                            if r_veh.get("exito"):
                                apply_mapping_logic("runt_vehiculo", r_veh["data"], cells)
                                apply_mapping_logic("rtm", r_veh["data"], cells)
                                apply_mapping_logic("soat", r_veh["data"], cells)
                                row_result["status"] = "FILA_COMPLETADA"
                            else:
                                cells[ExcelColumns.SOAT_ESTADO] = {"value": "Sin datos RUNT" if "no existe" in str(r_veh.get("mensaje","") or r_veh.get("error","")).lower() else "Error"}
                                row_result["status"] = "fallido"
                            
                            # RUNT Cit
                            if r_cit.get("exito"):
                                apply_mapping_logic("ciudadano", r_cit["data"], cells, extra={"categoria_labor": cat_labor})
                            else:
                                cells[ExcelColumns.LICENCIA_ESTADO] = {"value": "Sin datos RUNT" if "no existe" in str(r_cit.get("mensaje","") or r_cit.get("error","")).lower() else "Error"}
                                row_result["status"] = "fallido"

                            # SIMIT
                            if r_sim.get("exito"):
                                apply_mapping_logic("simit", r_sim["data"], cells)
                            else:
                                cells[ExcelColumns.SIMIT_ESTADO] = {"value": "Sin datos SIMIT"}
                                row_result["status"] = "fallido"

                            # Compliance calculation
                            c_rtm = str(cells.get(ExcelColumns.RTM_ESTADO, {}).get("value", "")).lower()
                            c_soat = str(cells.get(ExcelColumns.SOAT_ESTADO, {}).get("value", "")).lower()
                            c_lic = str(cells.get(ExcelColumns.LICENCIA_ESTADO, {}).get("value", "")).lower()
                            count = (1 if c_rtm == "vigente" else 0) + (1 if c_soat == "vigente" else 0) + (1 if c_lic == "vigente" else 0)
                            cells[ExcelColumns.CUMPLIMIENTO] = {"value": count}

                        except Exception as e:
                            logger.error(f"Row Query Error {cedula}: {e}")
                            row_result["status"] = "fallido"
                
                # Post-processing WebSocket update and Backend Status update
                if should_consult:
                    async with counter_lock:
                        processed_count += 1
                        if row_result["status"] == "FILA_COMPLETADA":
                            exitosos_count += 1
                        elif row_result["status"] == "fallido":
                            fallidos_count += 1
                        
                        live_results_map[my_index] = {
                            "index": my_index,
                            "status": row_result["status"], 
                            "cedula": cedula, 
                            "nombre": nombre
                        }
                        
                        progress = round((processed_count / total_filtrados) * 100, 1) if total_filtrados > 0 else 0
                        
                        if client_id:
                            msg = {
                                "index": my_index, "total": total_activos, "total_filtrados": total_filtrados,
                                "progress": progress, "cedula": cedula, "nombre": nombre, "status": row_result["status"]
                            }
                            await manager.send_personal_message(msg, client_id)
                            
                            # Save state periodically (every 5 rows or last row) to allow recovery
                            if processed_count % 5 == 0 or processed_count == total_filtrados:
                                # We sort the live results by index before saving
                                sorted_results = [live_results_map[k] for k in sorted(live_results_map.keys())]
                                await SeguridadVialService.save_task_status(client_id, identificacion, {
                                    "progress": progress,
                                    "exitosos": exitosos_count,
                                    "fallidos": fallidos_count,
                                    "processed": processed_count,
                                    "live_results": sorted_results
                                })
                
                results[row_idx] = row_result

            # 6. Execution Loop
            async with httpx.AsyncClient(timeout=30.0, follow_redirects=True, limits=httpx.Limits(max_connections=settings.BATCH_SIZE*2)) as client:
                tasks = [query_row(idx, client) for idx in all_valid_indices]
                await asyncio.gather(*tasks)

            # 7. Sequential Write to Excel (Safe)
            target_row = 10
            # Write Header for the new column and ensure style
            out_ws.cell(row=9, column=ExcelColumns.PROCESADO, value="PROCESADO")
            SeguridadVialService._copy_styles(out_ws.cell(row=9, column=ExcelColumns.FECHA_REVISION), out_ws.cell(row=9, column=ExcelColumns.PROCESADO))

            for row_idx in all_valid_indices:
                res = results.get(row_idx)
                if not res: continue
                
                # Copy original columns and basic info
                for col in range(1, in_ws.max_column + 1):
                    source_cell = in_ws.cell(row=row_idx, column=col)
                    target_cell = out_ws.cell(row=target_row, column=col)
                    target_cell.value = source_cell.value
                    if col <= 10 or col >= 22: SeguridadVialService._copy_styles(source_cell, target_cell)

                out_ws.cell(row=target_row, column=ExcelColumns.FECHA_REVISION, value=datetime.now().strftime("%d/%m/%Y"))

                # Apply Blueprint styles to the dynamic section
                for col in range(ExcelColumns.VIN, ExcelColumns.FECHA_REVISION + 1):
                    SeguridadVialService._copy_styles(out_ws.cell(row=10, column=col), out_ws.cell(row=target_row, column=col))
                    if col == ExcelColumns.RESTRICCIONES: out_ws.cell(row=target_row, column=col).fill = PatternFill(fill_type=None)

                # Write retrieved data
                row_cells = res.get("cells", {})
                for col, cell_data in row_cells.items():
                    cell = out_ws.cell(row=target_row, column=col, value=cell_data["value"])
                    if cell_data.get("is_normativa"):
                        cell.fill = PatternFill(start_color="FBD5B5", end_color="FBD5B5", fill_type="solid")
                    if cell_data.get("is_newline"):
                        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

                # Row height adjustment
                max_lines = 1
                for col in range(ExcelColumns.VIN, ExcelColumns.FECHA_REVISION + 1):
                    val = out_ws.cell(row=target_row, column=col).value
                    if val and isinstance(val, str): max_lines = max(max_lines, val.count('\n') + 1)
                out_ws.row_dimensions[target_row].height = max(30.75, max_lines * 15.5)
                
                # Column 'PROCESADO' Logic
                is_filtered = row_idx in filtered_indices
                if is_filtered:
                    proc_cell = out_ws.cell(row=target_row, column=ExcelColumns.PROCESADO, value="Consultado por CSR")
                    SeguridadVialService._copy_styles(out_ws.cell(row=10, column=ExcelColumns.FECHA_REVISION), proc_cell) # Copy style from metadata
                    if res.get("status") == "FILA_COMPLETADA":
                        proc_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid") # Vivid Green
                
                target_row += 1

            # Clean up extra template rows if any
            if out_ws.max_row >= target_row:
                rows_to_delete = out_ws.max_row - target_row + 1
                if rows_to_delete > 0: out_ws.delete_rows(target_row, rows_to_delete)

            # 8. Return Final Binary and Save to Disk
            output = io.BytesIO()
            out_wb.save(output)
            final_bytes = output.getvalue()
            
            if client_id:
                # Ensure the storage directory exists
                if not os.path.exists(settings.STORAGE_MASS_RESULTS):
                    os.makedirs(settings.STORAGE_MASS_RESULTS, exist_ok=True)
                    
                # Save result to disk
                res_path = SeguridadVialService._get_result_path(client_id)
                with open(res_path, "wb") as f:
                    f.write(final_bytes)
                
                # Update status to finished
                await SeguridadVialService.save_task_status(client_id, identificacion, {
                    "status": "PROCESO_FINALIZADO",
                    "progress": 100,
                    "exitosos": exitosos_count,
                    "fallidos": fallidos_count,
                    "result_ready": True,
                    "finish_time": datetime.now().isoformat(),
                    "live_results": [live_results_map[k] for k in sorted(live_results_map.keys())]
                })
                
                # Notify via WS
                logger.info(f"PROCESO COMPLETADO EXITOSAMENTE PARA CLIENTE: {client_id}")
                await manager.send_personal_message({
                    "status": "PROCESO_FINALIZADO",
                    "progress": 100,
                    "result_ready": True
                }, client_id)

        except Exception as e:
            logger.error(f"FATAL ERROR in procesar_excel background task: {str(e)}", exc_info=True)
            if client_id:
                try:
                    await SeguridadVialService.save_task_status(client_id, identificacion, {
                        "status": "error",
                        "error_detail": str(e),
                        "finish_time": datetime.now().isoformat()
                    })
                    await manager.send_personal_message({
                        "status": "error",
                        "error_detail": str(e)
                    }, client_id)
                except Exception as ex:
                    logger.error(f"Failed to save error status: {ex}")

        logger.info(f"Optimized report task finished. {total_filtrados} rows in {round(time.time() - start_time, 2)}s")
        return None
